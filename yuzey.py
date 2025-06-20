import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill, NamedStyle, Border, Side, Font
from openpyxl.utils import get_column_letter
import pandas as pd
from scipy.optimize import fsolve

sigma = 5.67e-8          # Stefan-Boltzmann sabiti (W/m^2 K^4)
eps   = 0.85             # Aviyonikler için ortalama epsilon değeri

# Excel dosyasını yükleme
excel_path = r"C:\Users\Delil\Desktop\VS_code\BTP\Table_A.4_&_Dimensions.xlsx"
try:
    table_a4   = pd.read_excel(excel_path, sheet_name="Table A.4")
    F_df       = pd.read_excel(excel_path, sheet_name="Dimensions")
    dimensions = pd.read_excel(excel_path, sheet_name="Dimensions")
except FileNotFoundError:
    raise FileNotFoundError(f"{excel_path} adresinde Excel bulunamadı. Tüm dosyalarınızın doğru adreste bulunduğundan emin olun.")


# Excel'deki Tablo A.4 verileri (100K-3000K arası)
temperatures, v, k, pr, beta = [table_a4.iloc[:, j].values
                                for j in range(5)]

# Excel'deki ısı kutusu verileri
element_names, A_upper, A_side, L_horizontal, L_vertical, Q_gen = [dimensions.iloc[:, j].values
                                                                   for j in range(6)]

# Sıcaklıkla değişen Rayleigh, Nusselt, Taşınım katsayısı ve bunlara bağlı sonuçlar için birer matris oluşturulur.
# Her matris (Tablonun satır sayısı  x  Eleman sayısı) boyutunda olup başlangıçta sıfırlardan oluşur.
# Örneğin "ra[50,3]" içerisinde 51. tablo satırı olan 400K için Ra değerini ve 4. elemanın boyutlarını barındırır.
num_elements     = len(element_names)
num_temperatures = len(temperatures)

# Excel'deki görme faktörü verileri
#F_elements = F_df.iloc[1:33, 6:39].to_numpy(dtype=float)
#F_ground_sky = F_df.iloc[1:33, 38:41].to_numpy(dtype=float)

F_elements   = pd.read_excel(excel_path, sheet_name="Dimensions", usecols="G:AL", header=None, skiprows=1, nrows=32).to_numpy()
F_ground_sky = pd.read_excel(excel_path, sheet_name="Dimensions", usecols="AM:AN", header=None, skiprows=1, nrows=32).to_numpy()

ra, nu, h, Q_convection_horizontal, Q_convection_vertical, Q_convection_outer = [
    np.zeros((num_temperatures, num_elements))
    for _ in range(6) # 6 : ra, nu, h, Q ...
]

# Sıcaklık tahminleri
T_ground = 323                              # Kabin zemininin yüzey sıcaklığı
T_satcom = 323                              # Satcom bölgesinin yüzey sıcaklığı
T_kabin  = 355.86                           # Kabin sıcaklığı için ilk tahmin (HAD sonucu [K])
T_guess  = np.full((num_elements), 200.0)  # Tahmini yüzey sıcaklığı (fsolve için matris halinde)
error_list = []
T_list     = []

# Enerji dengesi fonksiyonu
def energy_balance(T_surface):
    errors = np.zeros(num_elements)
    Q_rad_matrix = np.zeros((num_elements, num_elements))
    for d in range(num_elements):
        T_film_d = (T_surface[d] + T_kabin) / 2.0
        i = int(round(T_film_d - 100))
        i = max(0, i)
        i = min(num_temperatures-1, i)      
        
        # İlgili eleman için Rayleigh sayısı ------------------------------------------------------------------------#
        ra_current = abs((9.81 * beta[i] * (T_surface[d] - T_kabin) * (L_horizontal[d]**3.0) * pr[i]) / (v[i]**2.0)) #
        #------------------------------------------------------------------------------------------------------------#    
           
        # Üst yatay yüzey -------------------------------------------------------#       
        nu_horizontal = 0.028 * (ra_current**0.38)                               #
        h_horizontal = k[i] * nu_horizontal / L_horizontal[d]                    #
        Q_conv_horizontal = h_horizontal * A_upper[d] * (T_surface[d] - T_kabin) #
        #------------------------------------------------------------------------#       

        # Düşey yüzeyler ---------------------------------------------------#       
        nu_vertical = 0.35 * (ra_current**0.25)                             #
        h_vertical = k[i] * nu_vertical / L_vertical[d]                     #
        Q_conv_vertical = h_vertical * A_side[d] * (T_surface[d] - T_kabin) #
        #-------------------------------------------------------------------#       

        # Alt yatay yüzey ---------------------------------------------#       
        nu_outer = 0.39 * (ra_current**0.21)                           #
        h_outer = k[i] * nu_outer / L_vertical[d]                      #
        Q_conv_outer = h_outer * A_upper[d] * (T_surface[d] - T_kabin) #
        #--------------------------------------------------------------#
        
        for r in range(num_elements):
            A_r = 2 * A_upper[r] + A_side[r]
            for s in range(num_elements):
                if r != s:
                    Q_rad_matrix[r, s] = sigma * eps * A_r * F_elements[r , s] * (T_surface[r]**4.0 - T_surface[s]**4.0) 
        Q_box_rads = np.sum(Q_rad_matrix, axis=1)
        
        Q_rad_ground_sky = np.zeros(num_elements)
        for r in range(num_elements):
            A_i = 2 * A_upper[r] + A_side[r]
            Q_ground = sigma * eps * F_ground_sky[r, 0] * A_i * (T_surface[r]**4 - T_ground**4)
            Q_sky    = sigma * eps * F_ground_sky[r, 1] * A_i * (T_surface[r]**4 - T_satcom**4)        
            Q_rad_ground_sky[r] = Q_ground + Q_sky

        # Toplam ısı geçişi
        Q_total = Q_conv_horizontal + Q_conv_vertical + Q_conv_outer + Q_rad_ground_sky[d] + Q_box_rads[d]

        # Enerji dengesi & Hata
        errors[d] = Q_total - Q_gen[d]
        error_list.append(np.mean(errors))
        T_list.append(np.mean(T_surface))
        
    return errors

# Yüzey sıcaklığını bulan fonksiyon
T_surface_solutions = fsolve(energy_balance, T_guess)

# Sonuçları yazdırma
errors = energy_balance(T_surface_solutions)

# Excel çıktı
df_temps = pd.DataFrame({
    'Isı Kutusu': element_names,
    'Yüzey Sıcaklığı [°C]': T_surface_solutions - 273.15,
    'Hata': errors,
})

with pd.ExcelWriter(excel_path, mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
    df_temps.to_excel(writer, sheet_name='Yüzey Sıcaklıkarı', startcol=1, index=False)

# Excel Biçimlendirme
T_wb = load_workbook(excel_path)
T_ws = T_wb['Yüzey Sıcaklıkarı']

# Renkler
fill_colors = {
    "too_high_result":  '660000', # bordo
    "high_result":      'CC0000', # koyu kırmızı
    "regular_result":   'A9D08E', # yeşil
    "low_result":       '548235', # koyu yeşil
    "zero_result":      'D9D9D9', # gri
    "headers":          '1F4E79', # koyu mavi
    "toplam":           '9BC2E6'  # mavi
}
fills = {name: PatternFill(start_color=code, end_color=code, fill_type='solid') for name, code in fill_colors.items()}

# Kenarlıklar
thin_border = Border(
    left   =Side(style='thin', color='000000'),
    right  =Side(style='thin', color='000000'),
    top    =Side(style='thin', color='000000'),
    bottom =Side(style='thin', color='000000'),
)

# Hücre Biçimlendirme
def T_format(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            val = cell.value
            
            if isinstance(val, (int, float)):
                cell.font = Font(name='Arial', size=20, bold=False)
                cell.number_format = '0.0'
                if val > 300:
                    cell.fill = fills['too_high_results']
                    cell.font = Font(name='Arial', size=20, bold=False, color='FFFFFF')
                elif val > 100:
                    cell.fill = fills['high_result']
                    cell.font = Font(name='Arial', size=20, bold=False, color='FFFFFF')
                elif val > 50:
                    cell.fill = fills['regular_results']
                    cell.font = Font(name='Arial', size=20, bold=False, color='000000')
                elif val > 0.001:
                    cell.fill = fills['low_results']
                    cell.font = Font(name='Arial', size=20, bold=False, color='000000')
            elif isinstance(val, str):
                cell.fill = fills['headers']
                cell.font = Font(name='Arial', size=20, bold=False, color='000000')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

T_format(T_ws)

T_ws.sheet_view.zoomScale = 55
T_wb.save(excel_path)
print("🎨 Biçimlendirme tamamlandı.")

plt.plot(range(len(error_list)), error_list, label="Hata Değerleri")
plt.xlabel('İterasyonlar')
plt.ylabel('Hata Değerleri')
plt.title('Hata Takibi')
plt.grid(True)
plt.legend()
plt.show()